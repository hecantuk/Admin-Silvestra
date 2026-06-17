# ============================================================
#   "Jesús le dijo: Yo soy el camino, la verdad y la vida;
#    nadie viene al Padre sino por mí."
#                                         — Juan 14:6
#
#   "Todo lo puedo en Cristo que me fortalece."
#                                         — Filipenses 4:13
#
#   Sistema de Administración — Fraccionamiento Silvestra
#   Desarrollado con gratitud a Dios, que hace posibles
#   todas las cosas. A Él sea la honra y la gloria.
# ============================================================

from fastapi import FastAPI, HTTPException, Query, UploadFile, File, Depends
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, Response
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlmodel import Session, select, create_engine, SQLModel
from typing import Optional, List
from datetime import date, datetime, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
from pydantic import BaseModel
from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
import os, io, re, smtplib, base64
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders as _email_enc

from models import (Lote, Pago, Gasto, GastoArchivo, Usuario,
                    Proveedor, MovimientoBancario,
                    LecturaAgua, GastoReal, ProrrateoPorLote, Config, CuotaAnual, Descuento)

# ── Auth config ─────────────────────────────────────────────
SECRET_KEY  = os.environ.get("SECRET_KEY", "silvestra_dev_key_cambia_en_produccion")
ALGORITHM   = "HS256"
TOKEN_HOURS = 12

pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
bearer      = HTTPBearer(auto_error=False)

class LoginReq(BaseModel):
    username: str
    password: str

class CambiarPassReq(BaseModel):
    nueva: str

class ResetPassReq(BaseModel):
    user_key: str   # ej: M24L1

class ConfigEmailReq(BaseModel):
    smtp_from: str
    smtp_server: str = "smtp.office365.com"
    smtp_port: int = 587
    smtp_password: str = ""   # empty = keep existing password

class ConfigPlantillaReq(BaseModel):
    template: str
    asunto: str

class ConfigAutoReq(BaseModel):
    morosos: bool = False
    estado: bool = False
    dia: int = 5
    hora: str = "09:00"
    activo: bool = False

class EnviarCorreoReq(BaseModel):
    to: str
    subject: str
    body: str
    pdf_base64: str = ""
    pdf_filename: str = ""

def _make_token(user_id: int, rol: str, lote_id: Optional[int]) -> str:
    exp = datetime.utcnow() + timedelta(hours=TOKEN_HOURS)
    return jwt.encode({"sub": str(user_id), "rol": rol,
                       "lote_id": lote_id, "exp": exp},
                      SECRET_KEY, algorithm=ALGORITHM)

def _current_user(creds: HTTPAuthorizationCredentials = Depends(bearer),
                  session: Session = Depends(lambda: next(get_session()))):
    exc = HTTPException(status_code=401, detail="No autorizado")
    if not creds:
        raise exc
    try:
        payload = jwt.decode(creds.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        uid = payload.get("sub")
        if not uid:
            raise exc
    except JWTError:
        raise exc
    user = session.get(Usuario, int(uid))
    if not user or not user.activo:
        raise exc
    return user

def _admin_only(user: Usuario = Depends(_current_user)):
    if user.rol != "admin":
        raise HTTPException(status_code=403, detail="Solo administrador")
    return user

DATABASE_URL = os.environ.get("DATABASE_URL", "sqlite:///./silvestra.db")
# Railway usa postgres://, SQLAlchemy necesita postgresql://
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

connect_args = {"check_same_thread": False} if DATABASE_URL.startswith("sqlite") else {}
engine = create_engine(DATABASE_URL, echo=False, connect_args=connect_args)

# "El principio de la sabiduría es el temor del Señor." — Salmos 111:10

def get_session():
    with Session(engine) as session:
        yield session

# ── Config helpers ───────────────────────────────────────────
def _get_cfg(s: Session, clave: str, default: str = "") -> str:
    row = s.exec(select(Config).where(Config.clave == clave)).first()
    return row.valor if row else default

def _set_cfg(s: Session, clave: str, valor: str):
    row = s.exec(select(Config).where(Config.clave == clave)).first()
    if row:
        row.valor = valor
        s.add(row)
    else:
        s.add(Config(clave=clave, valor=valor))

# ── SMTP / email helpers ─────────────────────────────────────
_MESES = {
    "01": "enero", "02": "febrero", "03": "marzo", "04": "abril",
    "05": "mayo", "06": "junio", "07": "julio", "08": "agosto",
    "09": "septiembre", "10": "octubre", "11": "noviembre", "12": "diciembre",
}

def _send_smtp_msg(to: str, subject: str, body: str,
                   pdf_b64: str = "", pdf_name: str = ""):
    with Session(engine) as s:
        smtp_from     = _get_cfg(s, "smtp_from")
        smtp_server   = _get_cfg(s, "smtp_server", "smtp.office365.com")
        smtp_port_str = _get_cfg(s, "smtp_port", "587")
        smtp_password = _get_cfg(s, "smtp_password")
    if not smtp_from or not smtp_password:
        raise ValueError("SMTP no configurado. Ingresa correo y contraseña en Ajustes → Correo Outlook.")
    port = int(smtp_port_str)
    msg = MIMEMultipart()
    msg["From"] = smtp_from
    msg["To"] = to
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))
    if pdf_b64 and pdf_name:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(base64.b64decode(pdf_b64))
        _email_enc.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{pdf_name}"')
        msg.attach(part)
    with smtplib.SMTP(smtp_server, port, timeout=30) as srv:
        srv.ehlo()
        srv.starttls()
        srv.login(smtp_from, smtp_password)
        srv.send_message(msg)

def _build_server_email_body(template: str, lote: Lote,
                              saldo: float, estado: str, mes: str) -> str:
    y, m = mes.split("-")
    mes_label = f"{_MESES.get(m, m)} {y}"
    estatus = ("Al corriente" if estado == "corriente"
               else "Abonando a vencido" if estado == "abonando"
               else "Moroso")
    msg_est = ("Su cuenta se encuentra al corriente. ¡Gracias por su puntualidad!"
               if saldo <= 0
               else f"Tiene un saldo pendiente de ${saldo:,.0f}. Le invitamos a regularizar su situación.")
    if not template:
        return (f"Estimado/a {lote.propietario},\n\n"
                f"Le informamos el estado de su cuenta al mes de {mes_label}:\n\n"
                f"  Manzana {lote.manzana} — Lote {lote.numero}\n"
                f"  Cuota mensual: ${lote.cuota_cof:,.0f}\n"
                f"  Saldo pendiente: ${max(saldo,0):,.0f}\n"
                f"  Estatus: {estatus}\n\n"
                f"{msg_est}\n\n"
                f"Saludos,\nAdministración Silvestra-Canoas AC")
    return (template
            .replace("{{nombre}}", lote.propietario or "")
            .replace("{{lote}}", str(lote.numero))
            .replace("{{mza}}", str(lote.manzana))
            .replace("{{mes}}", mes_label)
            .replace("{{cuota}}", f"{lote.cuota_cof:,.0f}")
            .replace("{{saldo}}", f"{max(saldo, 0):,.0f}")
            .replace("{{estatus}}", estatus)
            .replace("{{mensajeEstatus}}", msg_est))

# ── APScheduler ──────────────────────────────────────────────
_scheduler = None

def _run_automatic_emails():
    """Executed by APScheduler at the configured time to send monthly emails."""
    print(f"[Scheduler] Iniciando envío automático {datetime.utcnow().isoformat()}")
    with Session(engine) as s:
        morosos_only = _get_cfg(s, "auto_morosos", "false") == "true"
        send_todos   = _get_cfg(s, "auto_estado",  "false") == "true"
        template     = _get_cfg(s, "email_template", "")
        asunto_tpl   = _get_cfg(s, "email_asunto", "Estado de cuenta {{mes}} - Silvestra")

    if not morosos_only and not send_todos:
        print("[Scheduler] Sin opciones activas, saliendo")
        return

    mes = datetime.utcnow().strftime("%Y-%m")
    with Session(engine) as s:
        lotes = s.exec(
            select(Lote).where(Lote.propietario != None, Lote.email != None)
        ).all()
        pagos = s.exec(
            select(Pago).where(Pago.mes_aplicado == mes, Pago.estado == "aprobado")
        ).all()
    pagos_lote: dict = {}
    for p in pagos:
        pagos_lote[p.lote_id] = pagos_lote.get(p.lote_id, 0.0) + p.importe

    y, m = mes.split("-")
    mes_label = f"{_MESES.get(m, m)} {y}"
    enviados = errores = 0
    for lote in lotes:
        if not lote.email:
            continue
        pagado = pagos_lote.get(lote.id, 0.0)
        saldo  = lote.cuota_cof - pagado
        estado = ("corriente" if saldo <= 0 else "abonando" if pagado > 0 else "moroso")
        if morosos_only and estado == "corriente":
            continue
        body    = _build_server_email_body(template, lote, saldo, estado, mes)
        subject = (asunto_tpl
                   .replace("{{mes}}", mes_label)
                   .replace("{{mza}}", str(lote.manzana))
                   .replace("{{lote}}", str(lote.numero)))
        try:
            _send_smtp_msg(lote.email, subject, body)
            enviados += 1
        except Exception as exc:
            print(f"[Scheduler] Error lote {lote.id}: {exc}")
            errores += 1
    print(f"[Scheduler] Completado: {enviados} enviados, {errores} errores")

def _reload_scheduler():
    global _scheduler
    if _scheduler is None:
        return
    try:
        _scheduler.remove_all_jobs()
        with Session(engine) as s:
            activo = _get_cfg(s, "auto_activo", "false") == "true"
            dia    = _get_cfg(s, "auto_dia", "5")
            hora   = _get_cfg(s, "auto_hora", "09:00")
        if activo:
            h, mi = hora.split(":")
            _scheduler.add_job(
                _run_automatic_emails,
                CronTrigger(day=dia, hour=int(h), minute=int(mi),
                            timezone="America/Monterrey"),
                id="email_auto", replace_existing=True,
            )
            print(f"[Scheduler] Job programado: día {dia} a las {hora} hora Monterrey")
        # Job mensual: actualizar cuotas el día 1 de cada mes a las 00:05
        _scheduler.add_job(
            lambda: _apply_cuotas_anuales(),
            CronTrigger(day=1, hour=0, minute=5, timezone="America/Monterrey"),
            id="cuotas_anuales", replace_existing=True,
        )
    except Exception as exc:
        print(f"[Scheduler] Error en _reload_scheduler: {exc}")

# "Y todo lo que hagan, de palabra o de obra,
#  háganlo en el nombre del Señor Jesús." — Colosenses 3:17
app = FastAPI(title="Silvestra Admin API", version="1.0.0")

# Servir el frontend
STATIC_DIR = os.path.join(os.path.dirname(__file__), "static")
if os.path.exists(STATIC_DIR):
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

@app.on_event("startup")
def on_startup():
    global _scheduler
    SQLModel.metadata.create_all(engine)
    # Add new Chequera columns to existing table if they don't exist yet
    from sqlalchemy import text as _text
    for _col_sql in [
        "ALTER TABLE movimientobancario ADD COLUMN nombre TEXT",
        "ALTER TABLE movimientobancario ADD COLUMN concepto TEXT",
        "ALTER TABLE movimientobancario ADD COLUMN importe DOUBLE PRECISION",
    ]:
        try:
            with engine.connect() as _c:
                _c.execute(_text(_col_sql))
                _c.commit()
        except Exception:
            pass
    try:
        _scheduler = BackgroundScheduler(timezone="America/Monterrey")
        _scheduler.start()
        _reload_scheduler()
        print("[Scheduler] Iniciado correctamente")
    except Exception as exc:
        print(f"[Scheduler] No se pudo iniciar: {exc}")
        _scheduler = None


# ─── ROOT ───────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
def root():
    index = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index):
        return FileResponse(index)
    return HTMLResponse("<h2>Copia tu index.html en la carpeta /static/</h2>")


# ─── AUTH ───────────────────────────────────────────────────
@app.post("/api/auth/login")
def login(req: LoginReq, session: Session = Depends(get_session)):
    user = None
    m = re.match(r'^M(\d+)L(\d+)$', req.username.strip().upper())
    if m:
        mza, num = int(m.group(1)), int(m.group(2))
        lote = session.exec(
            select(Lote).where(Lote.manzana == mza, Lote.numero == num)
        ).first()
        if lote:
            user = session.exec(
                select(Usuario).where(Usuario.lote_id == lote.id, Usuario.rol == "residente")
            ).first()
    else:
        user = session.exec(
            select(Usuario).where(Usuario.email == req.username.strip())
        ).first()

    if not user or not user.activo or not pwd_context.verify(req.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")

    token = _make_token(user.id, user.rol, user.lote_id)
    return {
        "token": token,
        "rol": user.rol,
        "lote_id": user.lote_id,
        "debe_cambiar_password": user.debe_cambiar_password,
    }

@app.post("/api/auth/cambiar-password")
def cambiar_password(req: CambiarPassReq,
                     user: Usuario = Depends(_current_user),
                     session: Session = Depends(get_session)):
    if len(req.nueva) < 6:
        raise HTTPException(status_code=400, detail="Mínimo 6 caracteres")
    db_user = session.get(Usuario, user.id)
    db_user.hashed_password = pwd_context.hash(req.nueva)
    db_user.debe_cambiar_password = False
    session.add(db_user)
    session.commit()
    return {"ok": True}

@app.post("/api/auth/reset-password")
def reset_password(req: ResetPassReq,
                   admin: Usuario = Depends(_admin_only),
                   session: Session = Depends(get_session)):
    m = re.match(r'^M(\d+)L(\d+)$', req.user_key.strip().upper())
    if not m:
        raise HTTPException(status_code=400, detail="Formato inválido (ej: M24L1)")
    mza, num = int(m.group(1)), int(m.group(2))
    lote = session.exec(
        select(Lote).where(Lote.manzana == mza, Lote.numero == num)
    ).first()
    if not lote:
        raise HTTPException(status_code=404, detail="Lote no encontrado")
    res_user = session.exec(
        select(Usuario).where(Usuario.lote_id == lote.id, Usuario.rol == "residente")
    ).first()
    if not res_user:
        raise HTTPException(status_code=404, detail="Usuario residente no encontrado")
    res_user.hashed_password = pwd_context.hash(f"silv{mza}{num}")
    res_user.debe_cambiar_password = True
    session.add(res_user)
    session.commit()
    return {"ok": True}

@app.get("/api/auth/residentes")
def get_residentes_auth(admin: Usuario = Depends(_admin_only),
                        session: Session = Depends(get_session)):
    residentes = session.exec(
        select(Usuario).where(Usuario.rol == "residente", Usuario.activo == True)
    ).all()
    result = []
    for u in residentes:
        lote = session.get(Lote, u.lote_id) if u.lote_id else None
        result.append({
            "user_key": u.email,
            "propietario": lote.propietario if lote else "—",
            "debe_cambiar_password": u.debe_cambiar_password,
        })
    return result

class NuevoAdminReq(BaseModel):
    email: str
    password: str

class CuotaAnualReq(BaseModel):
    anio: int
    min_m2: float
    max_m2: Optional[float] = None
    importe: float

class NuevoPagoReq(BaseModel):
    manzana: int
    lote_num: int
    fecha_pago: str   # "YYYY-MM-DD"
    importe: float
    concepto: str = "COF"
    referencia: str = ""
    notas: str = ""

@app.get("/api/auth/admins")
def get_admins(admin: Usuario = Depends(_admin_only),
               session: Session = Depends(get_session)):
    admins = session.exec(
        select(Usuario).where(Usuario.rol == "admin", Usuario.activo == True)
    ).all()
    return [{"id": u.id, "email": u.email} for u in admins]

@app.post("/api/auth/admins")
def crear_admin(req: NuevoAdminReq,
                admin: Usuario = Depends(_admin_only),
                session: Session = Depends(get_session)):
    if len(req.password) < 6:
        raise HTTPException(400, "Mínimo 6 caracteres")
    existe = session.exec(select(Usuario).where(Usuario.email == req.email)).first()
    if existe:
        raise HTTPException(400, "El usuario ya existe")
    session.add(Usuario(
        email=req.email,
        hashed_password=pwd_context.hash(req.password),
        rol="admin",
        debe_cambiar_password=False,
    ))
    session.commit()
    return {"ok": True}

@app.delete("/api/auth/admins/{uid}")
def eliminar_admin(uid: int,
                   admin: Usuario = Depends(_admin_only),
                   session: Session = Depends(get_session)):
    if admin.id == uid:
        raise HTTPException(400, "No puedes eliminarte a ti mismo")
    u = session.get(Usuario, uid)
    if not u or u.rol != "admin":
        raise HTTPException(404, "Admin no encontrado")
    u.activo = False
    session.add(u)
    session.commit()
    return {"ok": True}

@app.post("/api/auth/admins/{uid}/reset-password")
def reset_admin_password(uid: int, req: CambiarPassReq,
                         admin: Usuario = Depends(_admin_only),
                         session: Session = Depends(get_session)):
    if len(req.nueva) < 6:
        raise HTTPException(400, "Mínimo 6 caracteres")
    u = session.get(Usuario, uid)
    if not u or u.rol != "admin":
        raise HTTPException(404, "Admin no encontrado")
    u.hashed_password = pwd_context.hash(req.nueva)
    u.debe_cambiar_password = False
    u.activo = True
    session.add(u)
    session.commit()
    return {"ok": True}

@app.get("/api/auth/visores")
def get_visores(admin: Usuario = Depends(_admin_only),
                session: Session = Depends(get_session)):
    visores = session.exec(
        select(Usuario).where(Usuario.rol == "visor", Usuario.activo == True)
    ).all()
    return [{"id": u.id, "email": u.email} for u in visores]

@app.post("/api/auth/visores")
def crear_visor(req: NuevoAdminReq,
                admin: Usuario = Depends(_admin_only),
                session: Session = Depends(get_session)):
    if len(req.password) < 6:
        raise HTTPException(400, "Mínimo 6 caracteres")
    existe = session.exec(select(Usuario).where(Usuario.email == req.email)).first()
    if existe:
        raise HTTPException(400, "El usuario ya existe")
    session.add(Usuario(
        email=req.email,
        hashed_password=pwd_context.hash(req.password),
        rol="visor",
        debe_cambiar_password=False,
    ))
    session.commit()
    return {"ok": True}

@app.delete("/api/auth/visores/{uid}")
def eliminar_visor(uid: int,
                   admin: Usuario = Depends(_admin_only),
                   session: Session = Depends(get_session)):
    u = session.get(Usuario, uid)
    if not u or u.rol != "visor":
        raise HTTPException(404, "Visor no encontrado")
    u.activo = False
    session.add(u)
    session.commit()
    return {"ok": True}

# ─── LOTES ──────────────────────────────────────────────────
@app.get("/api/lotes")
def get_lotes(manzana: Optional[int] = None):
    with Session(engine) as s:
        q = select(Lote)
        if manzana:
            q = q.where(Lote.manzana == manzana)
        return s.exec(q).all()

@app.get("/api/lotes/{lote_id}")
def get_lote(lote_id: int):
    with Session(engine) as s:
        lote = s.get(Lote, lote_id)
        if not lote:
            raise HTTPException(404, "Lote no encontrado")
        return lote

@app.get("/api/lotes/{lote_id}/estado-cuenta")
def get_estado_cuenta(lote_id: int, user: Usuario = Depends(_current_user)):
    """Estado de cuenta completo: pagos y descuentos mensuales por lote."""
    with Session(engine) as s:
        lote = s.get(Lote, lote_id)
        if not lote:
            raise HTTPException(404, "Lote no encontrado")

        pagos = s.exec(
            select(Pago).where(Pago.lote_id == lote_id, Pago.estado == "aprobado")
        ).all()
        descuentos = s.exec(
            select(Descuento).where(Descuento.lote_id == lote_id)
        ).all()
        lecturas = s.exec(
            select(LecturaAgua).where(LecturaAgua.lote_id == lote_id)
        ).all()

        # Build monthly map
        by_mes: dict = {}
        for p in pagos:
            m = p.mes_aplicado or ""
            if m not in by_mes:
                by_mes[m] = {"pagoCof": 0, "pagoCov": 0, "fpago": str(p.fecha) if p.fecha else ""}
            if p.concepto == "COF":
                by_mes[m]["pagoCof"] += p.importe
                if p.fecha:
                    by_mes[m]["fpago"] = str(p.fecha)
            elif p.concepto == "COV":
                by_mes[m]["pagoCov"] += p.importe

        for d in descuentos:
            if d.anio:
                # Map annual discount to December of that year for display
                key = f"{d.anio}-12"
                if key not in by_mes:
                    by_mes[key] = {"pagoCof": 0, "pagoCov": 0, "fpago": ""}
                by_mes[key]["desc"] = by_mes[key].get("desc", 0) + d.importe

        lec_map: dict = {}
        for lec in lecturas:
            lec_map[lec.mes] = lec

        # Generate monthly rows from escrituracion to today
        hoy = date.today()
        inicio = lote.fecha_escrituracion or date(2020, 9, 1)
        cur = date(inicio.year, inicio.month, 1)
        meses_rows = []
        saldo_acum_cof = 0.0
        while cur <= date(hoy.year, hoy.month, 1):
            key = cur.strftime("%Y-%m")
            entry = by_mes.get(key, {})
            cof = lote.cuota_cof
            extra = 0.0
            pago_cof = entry.get("pagoCof", 0)
            desc = entry.get("desc", 0)
            fpago = entry.get("fpago", "")
            saldo_mes = extra + cof - pago_cof - desc
            saldo_acum_cof += saldo_mes
            lec = lec_map.get(key)
            meses_rows.append({
                "mes": key,
                "extra": extra,
                "cof": cof,
                "pagoCof": pago_cof,
                "desc": desc,
                "fpago": fpago,
                "saldoMes": round(saldo_mes, 2),
                "vIni": lec.lectura_inicial if lec else None,
                "vFin": lec.lectura_final if lec else None,
                "consumo": round((lec.lectura_final - lec.lectura_inicial), 2) if lec and lec.lectura_final and lec.lectura_inicial else None,
                "cov": round(lec.importe, 2) if lec else None,
                "instal": 0,
                "pagoCov": entry.get("pagoCov", 0),
            })
            cur = date(cur.year + (cur.month // 12), cur.month % 12 + 1, 1)

        total_pagado_cof = sum(p.importe for p in pagos if p.concepto == "COF")
        total_pagado_cov = sum(p.importe for p in pagos if p.concepto == "COV")
        total_desc = sum(d.importe for d in descuentos)
        total_cov_cargado = sum(lec.importe for lec in lecturas)
        n_meses = len(meses_rows)
        cargado_cof = lote.cuota_cof * n_meses
        saldo_cof = cargado_cof - total_pagado_cof - total_desc
        saldo_cov = total_cov_cargado - total_pagado_cov

        return {
            "lote_id": lote_id,
            "manzana": lote.manzana,
            "numero": lote.numero,
            "propietario": lote.propietario,
            "m2": lote.m2,
            "escrituracion": str(lote.fecha_escrituracion) if lote.fecha_escrituracion else None,
            "cuotaFija": lote.cuota_cof,
            "tarifaCOV": 5,
            "totales": {
                "cof": cargado_cof,
                "extra": 0,
                "pagoCof": total_pagado_cof,
                "desc": total_desc,
                "saldoCof": round(saldo_cof, 2),
                "cov": total_cov_cargado,
                "instal": 0,
                "pagoCov": total_pagado_cov,
            },
            "meses": meses_rows,
            "tiene_datos": len(pagos) > 0 or len(lecturas) > 0,
        }

@app.put("/api/lotes/{lote_id}")
def update_lote(lote_id: int, data: dict):
    with Session(engine) as s:
        lote = s.get(Lote, lote_id)
        if not lote:
            raise HTTPException(404, "Lote no encontrado")
        for k, v in data.items():
            setattr(lote, k, v)
        s.add(lote)
        s.commit()
        s.refresh(lote)
        return lote


# ─── PAGOS ──────────────────────────────────────────────────
@app.get("/api/pagos")
def get_pagos(estado: Optional[str] = None, mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(Pago)
        if estado:
            q = q.where(Pago.estado == estado)
        if mes:
            q = q.where(Pago.mes_aplicado == mes)
        return s.exec(q).all()

@app.post("/api/pagos")
def create_pago(pago: Pago):
    with Session(engine) as s:
        s.add(pago)
        s.commit()
        s.refresh(pago)
        return pago

@app.put("/api/pagos/{pago_id}/aprobar")
def aprobar_pago(pago_id: int):
    with Session(engine) as s:
        pago = s.get(Pago, pago_id)
        if not pago:
            raise HTTPException(404, "Pago no encontrado")
        pago.estado = "aprobado"
        s.add(pago)
        s.commit()
        return {"ok": True, "id": pago_id}

@app.put("/api/pagos/{pago_id}/revertir")
def revertir_pago(pago_id: int):
    with Session(engine) as s:
        pago = s.get(Pago, pago_id)
        if not pago:
            raise HTTPException(404, "Pago no encontrado")
        pago.estado = "por_aprobar"
        s.add(pago)
        s.commit()
        return {"ok": True, "id": pago_id}

@app.post("/api/pagos/nuevo")
def nuevo_pago(req: NuevoPagoReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        lote = s.exec(
            select(Lote).where(Lote.manzana == req.manzana, Lote.numero == req.lote_num)
        ).first()
        if not lote:
            raise HTTPException(404, "Lote no encontrado")
        fecha = date.fromisoformat(req.fecha_pago)
        mes_aplicado = req.fecha_pago[:7]
        pago = Pago(
            lote_id=lote.id,
            fecha_pago=fecha,
            importe=req.importe,
            mes_aplicado=mes_aplicado,
            concepto=req.concepto,
            referencia=req.referencia or None,
            notas=req.notas or None,
            estado="por_aprobar",
        )
        s.add(pago)
        s.commit()
        s.refresh(pago)
        return {
            "id": pago.id,
            "manzana": lote.manzana,
            "lote_num": lote.numero,
            "propietario": lote.propietario or "",
            "fecha_pago": str(pago.fecha_pago),
            "importe": pago.importe,
            "concepto": pago.concepto,
            "referencia": pago.referencia or "",
            "estado": pago.estado,
        }

@app.get("/api/pagos/lista")
def lista_pagos(estado: Optional[str] = None, mes: Optional[str] = None,
                admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        q = select(Pago, Lote).join(Lote, Pago.lote_id == Lote.id)
        if estado:
            q = q.where(Pago.estado == estado)
        if mes:
            q = q.where(Pago.mes_aplicado == mes)
        q = q.order_by(Pago.fecha_pago.desc())
        rows = s.exec(q).all()
        return [{
            "id": pago.id,
            "manzana": lote.manzana,
            "lote_num": lote.numero,
            "propietario": lote.propietario or "",
            "fecha_pago": str(pago.fecha_pago),
            "importe": pago.importe,
            "concepto": pago.concepto,
            "referencia": pago.referencia or "",
            "estado": pago.estado,
        } for pago, lote in rows]


# ─── GASTOS ──────────────────────────────────────────────────
@app.get("/api/gastos")
def get_gastos(mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(Gasto)
        if mes:
            q = q.where(Gasto.fecha.like(f"{mes}%"))
        return s.exec(q).all()

@app.post("/api/gastos")
def create_gasto(gasto: Gasto):
    with Session(engine) as s:
        s.add(gasto)
        s.commit()
        s.refresh(gasto)
        return gasto

@app.post("/api/gastos/{gid}/archivos")
async def subir_archivo_gasto(gid: int, archivo: UploadFile = File(...), admin: Usuario = Depends(_admin_only)):
    contenido = await archivo.read()
    if len(contenido) > 10_000_000:
        raise HTTPException(400, "Archivo muy grande (máx 10 MB)")
    b64 = base64.b64encode(contenido).decode()
    with Session(engine) as s:
        if not s.get(Gasto, gid):
            raise HTTPException(404, "Gasto no encontrado")
        a = GastoArchivo(gasto_id=gid, nombre=archivo.filename or "archivo",
                         tipo_mime=archivo.content_type or "application/octet-stream",
                         contenido_b64=b64)
        s.add(a); s.commit(); s.refresh(a)
        return {"ok": True, "id": a.id, "nombre": a.nombre}

@app.get("/api/gastos/{gid}/archivos")
def listar_archivos_gasto(gid: int, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        archivos = s.exec(select(GastoArchivo).where(GastoArchivo.gasto_id == gid)).all()
        return [{"id": a.id, "nombre": a.nombre, "tipo_mime": a.tipo_mime} for a in archivos]

@app.get("/api/archivos")
def listar_todos_archivos(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        archivos = s.exec(select(GastoArchivo)).all()
        return [{"id": a.id, "gasto_id": a.gasto_id, "nombre": a.nombre, "tipo_mime": a.tipo_mime} for a in archivos]

@app.get("/api/archivos/{aid}")
def ver_archivo(aid: int, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        a = s.get(GastoArchivo, aid)
        if not a:
            raise HTTPException(404, "Archivo no encontrado")
        contenido = base64.b64decode(a.contenido_b64)
        return Response(content=contenido, media_type=a.tipo_mime,
                        headers={"Content-Disposition": f'inline; filename="{a.nombre}"'})

@app.delete("/api/archivos/{aid}")
def eliminar_archivo(aid: int, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        a = s.get(GastoArchivo, aid)
        if not a:
            raise HTTPException(404)
        s.delete(a); s.commit()
        return {"ok": True}


# ─── REPORTE / SALDOS ────────────────────────────────────────
@app.get("/api/reporte/saldos")
def reporte_saldos(mes: str = "2026-05"):
    """Concentrado de cobranza acumulado: saldo real de cada lote al mes dado."""
    with Session(engine) as s:
        lotes = s.exec(select(Lote).where(Lote.propietario != None)).all()
        resultado = []
        for lote in lotes:
            # Pagos acumulados aprobados hasta el mes dado (COF y COV por separado)
            pagos = s.exec(
                select(Pago).where(
                    Pago.lote_id == lote.id,
                    Pago.estado == "aprobado",
                    Pago.mes_aplicado <= mes,
                )
            ).all()
            pagado_cof = sum(p.importe for p in pagos if p.concepto == "COF")
            pagado_cov = sum(p.importe for p in pagos if p.concepto == "COV")

            # Descuentos acumulados
            descuentos = s.exec(
                select(Descuento).where(Descuento.lote_id == lote.id)
            ).all()
            desc_cof = sum(d.importe for d in descuentos if d.concepto == "COF")
            desc_cov = sum(d.importe for d in descuentos if d.concepto == "COV")
            desc_total = desc_cof + desc_cov

            # COF acumulado: cuota × meses desde escrituración hasta mes dado
            anio_mes = mes.split("-")
            mes_hasta = date(int(anio_mes[0]), int(anio_mes[1]), 1)
            if lote.fecha_escrituracion:
                mes_inicio = date(lote.fecha_escrituracion.year, lote.fecha_escrituracion.month, 1)
            else:
                mes_inicio = date(2020, 9, 1)  # fallback: primer registro en el sistema
            meses_transcurridos = max(0, (mes_hasta.year - mes_inicio.year) * 12 + (mes_hasta.month - mes_inicio.month) + 1)
            cargado_cof = lote.cuota_cof * meses_transcurridos

            # COV cargado: suma de lecturas de agua aprobadas
            cov_cargado = s.exec(
                select(LecturaAgua).where(LecturaAgua.lote_id == lote.id, LecturaAgua.mes <= mes)
            ).all()
            cargado_cov = sum(l.importe for l in cov_cargado)

            saldo_cof = cargado_cof - pagado_cof - desc_cof
            saldo_cov = cargado_cov - pagado_cov - desc_cov
            saldo_total = saldo_cof + saldo_cov

            pagado_mes = sum(p.importe for p in pagos if p.mes_aplicado == mes)
            meses_vencidos = max(0, round(saldo_cof / lote.cuota_cof)) if saldo_cof > 0 else 0

            resultado.append({
                "lote_id": lote.id,
                "manzana": lote.manzana,
                "numero": lote.numero,
                "propietario": lote.propietario,
                "cuota": lote.cuota_cof,
                "cargado_cof": round(cargado_cof, 2),
                "pagado_cof": round(pagado_cof, 2),
                "pagado_cov": round(pagado_cov, 2),
                "descuentos": round(desc_total, 2),
                "desc_cof": round(desc_cof, 2),
                "desc_cov": round(desc_cov, 2),
                "saldo_cof": round(saldo_cof, 2),
                "saldo_cov": round(saldo_cov, 2),
                "saldo": round(saldo_total, 2),
                "pagado": round(pagado_cof + pagado_cov, 2),
                "meses_transcurridos": meses_transcurridos,
                "meses_vencidos": meses_vencidos,
                "estado": "corriente" if saldo_total <= 0 else ("abonando" if pagado_mes > 0 else "moroso"),
            })
        total = len(resultado)
        corriente = sum(1 for r in resultado if r["estado"] == "corriente")
        morosos = sum(1 for r in resultado if r["estado"] == "moroso")
        return {
            "mes": mes,
            "total_lotes": total,
            "corriente": corriente,
            "morosos": morosos,
            "abonando": total - corriente - morosos,
            "pct_corriente": round(corriente / total * 100, 1) if total else 0,
            "pct_morosos": round(morosos / total * 100, 1) if total else 0,
            "detalle": resultado,
        }



# ─── DESCUENTOS ───────────────────────────────────────────────
class DescuentoReq(SQLModel):
    lote_id: int
    tipo: str = "descuento"
    concepto: str = "COF"
    importe: float
    anio: Optional[int] = None
    notas: Optional[str] = None
    fecha: Optional[str] = None

@app.get("/api/descuentos/resumen")
def get_descuentos_resumen(admin: Usuario = Depends(_admin_only)):
    """KPI summary for Descuentos page."""
    with Session(engine) as s:
        rows = s.exec(select(Descuento)).all()
        total = sum(d.importe for d in rows)
        lotes_ids = set(d.lote_id for d in rows)
        return {
            "total_descuentos": round(total, 2),
            "lotes_con_descuento": len(lotes_ids),
            "promedio": round(total / len(lotes_ids), 2) if lotes_ids else 0,
        }

@app.get("/api/descuentos")
def get_descuentos(lote_id: Optional[int] = None, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        q = select(Descuento, Lote).join(Lote, Descuento.lote_id == Lote.id)
        if lote_id:
            q = q.where(Descuento.lote_id == lote_id)
        rows = s.exec(q).all()
        return [{
            "id": d.id,
            "lote_id": d.lote_id,
            "manzana": l.manzana,
            "numero": l.numero,
            "propietario": l.propietario or "",
            "tipo": d.tipo,
            "concepto": d.concepto,
            "importe": d.importe,
            "anio": d.anio,
            "notas": d.notas,
            "fecha": str(d.fecha),
        } for d, l in rows]

@app.post("/api/descuentos")
def crear_descuento(req: DescuentoReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        lote = s.get(Lote, req.lote_id)
        if not lote:
            raise HTTPException(404, "Lote no encontrado")
        d = Descuento(
            lote_id=req.lote_id,
            tipo=req.tipo,
            concepto=req.concepto,
            importe=req.importe,
            anio=req.anio,
            notas=req.notas or None,
            fecha=date.fromisoformat(req.fecha) if req.fecha else date.today(),
        )
        s.add(d)
        s.commit()
        s.refresh(d)
        return {"id": d.id, "ok": True}

@app.delete("/api/descuentos/{did}")
def eliminar_descuento(did: int, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        d = s.get(Descuento, did)
        if not d:
            raise HTTPException(404, "Descuento no encontrado")
        s.delete(d)
        s.commit()
        return {"ok": True}


# ─── STATS DASHBOARD ─────────────────────────────────────────
@app.get("/api/stats")
def get_stats(mes: str = "2026-05"):
    with Session(engine) as s:
        lotes = s.exec(select(Lote).where(Lote.propietario != None)).all()
        pagos_mes = s.exec(
            select(Pago).where(Pago.mes_aplicado == mes, Pago.estado == "aprobado")
        ).all()
        total_cuotas = sum(l.cuota_cof for l in lotes)
        total_cobrado = sum(p.importe for p in pagos_mes)
        return {
            "lotes_vendidos": len(lotes),
            "total_cuotas_mes": total_cuotas,
            "total_cobrado": total_cobrado,
            "pendiente": total_cuotas - total_cobrado,
            "pct_cobrado": round(total_cobrado / total_cuotas * 100, 1) if total_cuotas else 0,
        }


# "Fíate del Señor de todo tu corazón, y no te apoyes en
#  tu propio entendimiento. Reconócelo en todos tus caminos,
#  y Él enderezará tus veredas." — Proverbios 3:5-6
M2_TOTALES = 299174.17
CONCEPTOS = [
    "LUZ","REDES DE AGUA","CUOTAS DE SEGUROS","RESIDENTFY","INTERNET",
    "MAQ Y EQ","MATERIALES","MTTO FRACCIONAMIENTO","MTTO INFRAESTRUCTURA",
    "GASOLINA","SEGURIDAD","ISR RETENIDO","IVA RETENIDO",
    "COMISION BANCO","IVA GASTOS","NOMINA",
]


# ─── PROVEEDORES ─────────────────────────────────────────────
@app.get("/api/proveedores")
def get_proveedores():
    with Session(engine) as s:
        return s.exec(select(Proveedor)).all()

@app.post("/api/proveedores")
def create_proveedor(p: Proveedor):
    with Session(engine) as s:
        s.add(p); s.commit(); s.refresh(p)
        return p

@app.put("/api/proveedores/{pid}")
def update_proveedor(pid: int, data: dict):
    with Session(engine) as s:
        p = s.get(Proveedor, pid)
        if not p: raise HTTPException(404, "Proveedor no encontrado")
        for k, v in data.items(): setattr(p, k, v)
        s.add(p); s.commit(); s.refresh(p)
        return p

@app.delete("/api/proveedores/{pid}")
def delete_proveedor(pid: int):
    with Session(engine) as s:
        p = s.get(Proveedor, pid)
        if not p: raise HTTPException(404)
        s.delete(p); s.commit()
        return {"ok": True}


# ─── MÓDULO 1: BANCO ─────────────────────────────────────────
@app.post("/api/admin/limpiar_chequera")
async def limpiar_chequera(admin: Usuario = Depends(_admin_only)):
    from sqlalchemy import text as _text2
    with engine.connect() as conn:
        conn.execute(_text2("DELETE FROM movimientobancario"))
        conn.commit()
    return {"ok": True}

@app.post("/api/admin/limpiar_todo")
async def limpiar_todo(admin: Usuario = Depends(_admin_only)):
    """Borra todos los datos operativos. Conserva usuarios, config y lotes."""
    import json as _json, os as _os
    from sqlalchemy import text as _text2
    tables = [
        "movimientobancario", "pago", "gasto", "gastoarchivo",
        "descuento", "lecturaagua", "gastoreal", "prorrateoporldote",
    ]
    with engine.connect() as conn:
        for t in tables:
            try:
                conn.execute(_text2(f"DELETE FROM {t}"))
            except Exception:
                pass
        conn.commit()
    # Vaciar silvestra_data.json
    json_path = _os.path.join(_os.path.dirname(__file__), "static", "silvestra_data.json")
    with open(json_path, "w") as f:
        f.write("[]")
    return {"ok": True}

@app.get("/api/banco/movimientos")
def get_movimientos(mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(MovimientoBancario)
        if mes:
            q = q.where(MovimientoBancario.fecha.like(f"{mes}%"))
        movs = s.exec(q.order_by(MovimientoBancario.fecha.desc())).all()
        total_abonos = sum(m.abono for m in movs)
        total_cargos = sum(m.cargo for m in movs)
        return {"movimientos": movs, "total_abonos": total_abonos, "total_cargos": total_cargos}

@app.post("/api/banco/importar")
async def importar_banco(file: UploadFile = File(...)):
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    # Detectar fila de encabezado buscando palabras clave
    header_row = 0
    for i, row in enumerate(rows[:10]):
        vals = [str(v).lower() if v else "" for v in row]
        if any("fecha" in v for v in vals) and (any("descripci" in v for v in vals) or any("nombre" in v or "cheque" in v for v in vals)):
            header_row = i
            break
    headers = [str(v).lower().strip() if v else "" for v in rows[header_row]]
    col = {}
    for i, h in enumerate(headers):
        if "fecha" in h: col["fecha"] = i
        elif "nombre" in h: col["nombre"] = i
        elif "concepto" in h: col["concepto"] = i
        elif "descripci" in h: col["desc"] = i
        elif "importe" in h: col["importe"] = i
        elif "ret" in h and "isr" in h: col["ret_isr"] = i
        elif "ret" in h and "iva" in h: col["ret_iva"] = i
        elif "iva" in h: col["iva"] = i
        elif "debe" in h: col["debe"] = i
        elif "haber" in h: col["haber"] = i
        elif "cheque" in h: col["cheque"] = i
        elif "cargo" in h or "retiro" in h or "debito" in h: col["cargo"] = i
        elif "abono" in h or "deposito" in h or "credito" in h: col["abono"] = i
        elif "referen" in h or "folio" in h: col["ref"] = i

    # Detect 12-column Chequera format (has NOMBRE and/or CONCEPTO columns)
    is_chequera_fmt = "nombre" in col or "concepto" in col or "cheque" in col

    importados = 0
    with Session(engine) as s:
        proveedores = s.exec(select(Proveedor)).all()
        lotes = s.exec(select(Lote).where(Lote.propietario != None)).all()
        for row in rows[header_row + 1:]:
            if not row: continue
            fecha_raw = row[col.get("fecha", 2 if is_chequera_fmt else 0)]
            if not fecha_raw: continue
            if hasattr(fecha_raw, "date"):
                fecha = fecha_raw.date()
            else:
                try: fecha = date.fromisoformat(str(fecha_raw)[:10])
                except:
                    # Try Excel serial date
                    try:
                        from datetime import date as _date, timedelta as _td
                        fecha = _date(1899, 12, 30) + _td(days=int(float(str(fecha_raw))))
                    except: continue

            if is_chequera_fmt:
                nombre_v = str(row[col["nombre"]] or "") if "nombre" in col else ""
                concepto_v = str(row[col["concepto"]] or "") if "concepto" in col else ""
                desc = nombre_v or concepto_v or "Sin descripción"
                importe_v = float(row[col["importe"]] or 0) if "importe" in col and row[col["importe"]] else None
                iva_v = float(row[col["iva"]] or 0) if "iva" in col and row[col["iva"]] else 0.0
                ret_isr_v = float(row[col["ret_isr"]] or 0) if "ret_isr" in col and row[col["ret_isr"]] else 0.0
                ret_iva_v = float(row[col["ret_iva"]] or 0) if "ret_iva" in col and row[col["ret_iva"]] else 0.0
                cargo = float(row[col["debe"]] or 0) if "debe" in col and row[col["debe"]] else 0.0
                abono = float(row[col["haber"]] or 0) if "haber" in col and row[col["haber"]] else 0.0
                ref = str(row[col["cheque"]] or "") if "cheque" in col and row[col["cheque"]] else None
            else:
                nombre_v = None; concepto_v = None; importe_v = None
                iva_v = 0.0; ret_isr_v = 0.0; ret_iva_v = 0.0
                desc = str(row[col.get("desc", 1)] or "")
                cargo = float(row[col.get("cargo", 2)] or 0)
                abono = float(row[col.get("abono", 3)] or 0)
                ref = str(row[col.get("ref", 4)] or "") if "ref" in col else None

            # Clasificar automáticamente
            tipo = "sin_clasificar"; lote_id = None; prov_id = None
            search_text = (nombre_v or desc).lower()
            for lote in lotes:
                if lote.propietario:
                    apellido = lote.propietario.split()[-1].lower()
                    if len(apellido) > 3 and apellido in search_text:
                        tipo = "ingreso_colono"; lote_id = lote.id; break
            if tipo == "sin_clasificar":
                for prov in proveedores:
                    if prov.nombre.split()[0].lower() in search_text:
                        tipo = "pago_proveedor"; prov_id = prov.id; break

            mov = MovimientoBancario(
                fecha=fecha, descripcion=desc, nombre=nombre_v, concepto=concepto_v,
                importe=importe_v, cargo=cargo, abono=abono,
                referencia=ref, tipo=tipo, lote_id=lote_id, proveedor_id=prov_id,
                iva_ret=iva_v or ret_iva_v, isr_ret=ret_isr_v,
                tiene_iva_ret=bool(iva_v or ret_iva_v), tiene_isr_ret=bool(ret_isr_v),
            )
            s.add(mov)
            importados += 1
        s.commit()
    return {"importados": importados}

@app.put("/api/banco/movimientos/{mid}")
def update_movimiento(mid: int, data: dict):
    with Session(engine) as s:
        m = s.get(MovimientoBancario, mid)
        if not m: raise HTTPException(404)
        for k, v in data.items(): setattr(m, k, v)
        s.add(m); s.commit(); s.refresh(m)
        return m

@app.delete("/api/banco/movimientos/{mid}")
def delete_movimiento(mid: int):
    with Session(engine) as s:
        m = s.get(MovimientoBancario, mid)
        if not m: raise HTTPException(404)
        s.delete(m); s.commit()
        return {"ok": True}


# ─── MÓDULO 2: DIOT ──────────────────────────────────────────
@app.get("/api/diot")
def get_diot(mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(MovimientoBancario).where(MovimientoBancario.tipo == "pago_proveedor")
        if mes:
            q = q.where(MovimientoBancario.fecha.like(f"{mes}%"))
        movs = s.exec(q).all()
        provs = {p.id: p for p in s.exec(select(Proveedor)).all()}
        rows = []
        for m in movs:
            p = provs.get(m.proveedor_id)
            rows.append({
                "id": m.id,
                "fecha": str(m.fecha),
                "proveedor": p.nombre if p else m.descripcion,
                "rfc": p.rfc if p else None,
                "tipo_persona": p.tipo_persona if p else None,
                "importe": m.cargo,
                "iva_ret": m.iva_ret,
                "isr_ret": m.isr_ret,
            })
        return {"mes": mes, "filas": rows,
                "total_importe": sum(r["importe"] for r in rows),
                "total_iva_ret": sum(r["iva_ret"] for r in rows),
                "total_isr_ret": sum(r["isr_ret"] for r in rows)}


# ─── MÓDULO 3: AGUA ──────────────────────────────────────────
@app.get("/api/agua/lecturas")
def get_lecturas(mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(LecturaAgua)
        if mes: q = q.where(LecturaAgua.mes == mes)
        lecturas = s.exec(q).all()
        lotes = {l.id: l for l in s.exec(select(Lote).where(Lote.propietario != None)).all()}
        result = []
        for lec in lecturas:
            lote = lotes.get(lec.lote_id)
            result.append({**lec.dict(),
                           "propietario": lote.propietario if lote else None,
                           "manzana": lote.manzana if lote else None,
                           "numero": lote.numero if lote else None})
        return result

@app.post("/api/agua/lecturas")
def create_lectura(lec: LecturaAgua):
    lec.consumo_m3 = lec.lectura_actual - lec.lectura_anterior
    lec.importe = round(lec.consumo_m3 * lec.tarifa_por_m3, 2)
    with Session(engine) as s:
        s.add(lec); s.commit(); s.refresh(lec)
        return lec

@app.put("/api/agua/lecturas/{lid}")
def update_lectura(lid: int, data: dict):
    with Session(engine) as s:
        lec = s.get(LecturaAgua, lid)
        if not lec: raise HTTPException(404)
        for k, v in data.items(): setattr(lec, k, v)
        lec.consumo_m3 = lec.lectura_actual - lec.lectura_anterior
        lec.importe = round(lec.consumo_m3 * lec.tarifa_por_m3, 2)
        s.add(lec); s.commit(); s.refresh(lec)
        return lec

@app.delete("/api/agua/lecturas/{lid}")
def delete_lectura(lid: int):
    with Session(engine) as s:
        lec = s.get(LecturaAgua, lid)
        if not lec: raise HTTPException(404)
        s.delete(lec); s.commit()
        return {"ok": True}


@app.post("/api/agua/importar")
async def importar_agua(file: UploadFile = File(...), mes: str = "2026-05", tarifa: float = 15.0):
    """
    Importa lecturas de agua desde Excel.
    Columnas esperadas (en cualquier orden, nombres flexibles):
      manzana, lote/numero, lectura_anterior, lectura_actual
    También acepta: mza, no, lote_ant, lect_ant, ant, anterior, actual, lect_act
    """
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Detectar fila de encabezado
    header_row = 0
    for i, row in enumerate(rows[:10]):
        vals = [str(v).lower().strip() if v else "" for v in row]
        if any(k in " ".join(vals) for k in ["manzana","mza","lote","numero"]):
            header_row = i
            break

    headers = [str(v).lower().strip() if v else "" for v in rows[header_row]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(k in h for k in keywords):
                return i
        return None

    col_mza  = find_col(["manzana","mza"])
    col_lote = find_col(["lote","numero","no.","num"])
    col_ant  = find_col(["anterior","ant","lect_ant","lectura_ant","lectura ant"])
    col_act  = find_col(["actual","act","lect_act","lectura_act","lectura act","nueva"])

    if col_lote is None or col_act is None:
        raise HTTPException(400, "No se encontraron columnas de Lote y/o Lectura Actual. "
                                 "Asegúrate de que el Excel tenga columnas: Manzana, Lote, Lectura Anterior, Lectura Actual")

    importados = 0
    errores = []
    with Session(engine) as s:
        lotes = s.exec(select(Lote)).all()
        # índice: (manzana, numero) → lote
        lotes_idx = {(l.manzana, l.numero): l for l in lotes}

        for i, row in enumerate(rows[header_row + 1:], start=header_row + 2):
            if not row or all(v is None for v in row):
                continue
            try:
                mza_val  = int(float(str(row[col_mza]).strip())) if col_mza is not None and row[col_mza] else None
                lote_val = int(float(str(row[col_lote]).strip())) if row[col_lote] else None
                ant_val  = float(str(row[col_ant]).strip().replace(",", "")) if col_ant is not None and row[col_ant] else 0.0
                act_val  = float(str(row[col_act]).strip().replace(",", "")) if row[col_act] else None

                if lote_val is None or act_val is None:
                    continue

                # Buscar lote: primero con manzana, luego solo por número
                lote = None
                if mza_val:
                    lote = lotes_idx.get((mza_val, lote_val))
                if not lote:
                    candidatos = [l for l in lotes if l.numero == lote_val]
                    if len(candidatos) == 1:
                        lote = candidatos[0]

                if not lote:
                    errores.append(f"Fila {i}: lote {mza_val}-{lote_val} no encontrado")
                    continue

                consumo = max(0.0, act_val - ant_val)
                importe = round(consumo * tarifa, 2)

                # Upsert: si ya existe lectura para este lote+mes la actualiza
                existing = s.exec(
                    select(LecturaAgua).where(LecturaAgua.lote_id == lote.id, LecturaAgua.mes == mes)
                ).first()

                if existing:
                    existing.lectura_anterior = ant_val
                    existing.lectura_actual   = act_val
                    existing.consumo_m3       = consumo
                    existing.tarifa_por_m3    = tarifa
                    existing.importe          = importe
                    s.add(existing)
                else:
                    s.add(LecturaAgua(
                        lote_id=lote.id, mes=mes,
                        lectura_anterior=ant_val, lectura_actual=act_val,
                        consumo_m3=consumo, tarifa_por_m3=tarifa, importe=importe,
                    ))
                importados += 1
            except Exception as e:
                errores.append(f"Fila {i}: {e}")

        s.commit()

    return {"importados": importados, "errores": errores}


@app.get("/api/agua/reporte")
def reporte_agua(mes: str = "2026-05"):
    with Session(engine) as s:
        lecturas = s.exec(select(LecturaAgua).where(LecturaAgua.mes == mes)).all()
        mes_anterior = mes[:5] + str(int(mes[5:]) - 1).zfill(2) if int(mes[5:]) > 1 else str(int(mes[:4]) - 1) + "-12"
        lects_ant = {l.lote_id: l for l in s.exec(select(LecturaAgua).where(LecturaAgua.mes == mes_anterior)).all()}
        lotes = {l.id: l for l in s.exec(select(Lote)).all()}
        result = []
        for lec in lecturas:
            lote = lotes.get(lec.lote_id)
            ant = lects_ant.get(lec.lote_id)
            result.append({
                "lote_id": lec.lote_id,
                "propietario": lote.propietario if lote else None,
                "manzana": lote.manzana if lote else None,
                "numero": lote.numero if lote else None,
                "consumo_actual": lec.consumo_m3,
                "consumo_anterior": ant.consumo_m3 if ant else None,
                "variacion": round(lec.consumo_m3 - ant.consumo_m3, 2) if ant else None,
                "importe": lec.importe,
            })
        return {"mes": mes, "lecturas": result,
                "total_consumo": sum(r["consumo_actual"] for r in result),
                "total_importe": sum(r["importe"] for r in result)}


# ─── MÓDULO 4: PRORRATEO ─────────────────────────────────────
@app.get("/api/prorrateo/conceptos")
def get_conceptos():
    return CONCEPTOS

@app.get("/api/prorrateo/gastos-reales")
def get_gastos_reales(mes: Optional[str] = None):
    with Session(engine) as s:
        q = select(GastoReal)
        if mes: q = q.where(GastoReal.mes == mes)
        return s.exec(q).all()

@app.post("/api/prorrateo/gastos-reales")
def upsert_gasto_real(gr: GastoReal):
    with Session(engine) as s:
        existing = s.exec(
            select(GastoReal).where(GastoReal.mes == gr.mes, GastoReal.concepto == gr.concepto)
        ).first()
        if existing:
            existing.importe = gr.importe
            existing.notas = gr.notas
            s.add(existing); s.commit(); s.refresh(existing)
            return existing
        s.add(gr); s.commit(); s.refresh(gr)
        return gr

@app.post("/api/prorrateo/calcular/{mes}")
def calcular_prorrateo(mes: str):
    with Session(engine) as s:
        gastos = s.exec(select(GastoReal).where(GastoReal.mes == mes)).all()
        if not gastos:
            raise HTTPException(400, "No hay gastos reales para este mes")
        lotes = s.exec(select(Lote)).all()
        # Borrar prorrateo anterior del mes
        s.exec(ProrrateoPorLote.__table__.delete().where(ProrrateoPorLote.mes == mes))
        s.commit()
        registros = 0
        for gasto in gastos:
            costo_m2 = gasto.importe / M2_TOTALES
            for lote in lotes:
                importe = round(costo_m2 * lote.m2, 4)
                es_campestre = lote.propietario is None
                s.add(ProrrateoPorLote(
                    lote_id=lote.id, mes=mes,
                    concepto=gasto.concepto,
                    importe_prorrateado=importe,
                    es_campestre=es_campestre,
                ))
                registros += 1
        s.commit()
        # Resumen
        todos = s.exec(select(ProrrateoPorLote).where(ProrrateoPorLote.mes == mes)).all()
        total_asociados = sum(p.importe_prorrateado for p in todos if not p.es_campestre)
        total_campestre = sum(p.importe_prorrateado for p in todos if p.es_campestre)
        return {"mes": mes, "registros": registros,
                "total_asociados": round(total_asociados, 2),
                "total_campestre": round(total_campestre, 2),
                "total_general": round(total_asociados + total_campestre, 2)}

@app.get("/api/prorrateo/resumen/{mes}")
def resumen_prorrateo(mes: str):
    with Session(engine) as s:
        rows = s.exec(select(ProrrateoPorLote).where(ProrrateoPorLote.mes == mes)).all()
        if not rows:
            return {"mes": mes, "calculado": False}
        gastos = {g.concepto: g.importe for g in s.exec(select(GastoReal).where(GastoReal.mes == mes)).all()}
        por_concepto = {}
        for r in rows:
            if r.concepto not in por_concepto:
                por_concepto[r.concepto] = {"concepto": r.concepto,
                                             "gasto_total": gastos.get(r.concepto, 0),
                                             "asociados": 0, "campestre": 0}
            if r.es_campestre: por_concepto[r.concepto]["campestre"] += r.importe_prorrateado
            else: por_concepto[r.concepto]["asociados"] += r.importe_prorrateado
        conceptos_list = list(por_concepto.values())
        for c in conceptos_list:
            c["asociados"] = round(c["asociados"], 2)
            c["campestre"] = round(c["campestre"], 2)
        return {"mes": mes, "calculado": True,
                "conceptos": conceptos_list,
                "total_asociados": round(sum(c["asociados"] for c in conceptos_list), 2),
                "total_campestre": round(sum(c["campestre"] for c in conceptos_list), 2)}

@app.get("/api/prorrateo/por-lote/{mes}")
def prorrateo_por_lote(mes: str, solo_campestre: bool = False):
    with Session(engine) as s:
        q = select(ProrrateoPorLote).where(ProrrateoPorLote.mes == mes)
        if solo_campestre: q = q.where(ProrrateoPorLote.es_campestre == True)
        rows = s.exec(q).all()
        lotes = {l.id: l for l in s.exec(select(Lote)).all()}
        por_lote = {}
        for r in rows:
            if r.lote_id not in por_lote:
                lote = lotes.get(r.lote_id)
                por_lote[r.lote_id] = {"lote_id": r.lote_id,
                                        "manzana": lote.manzana if lote else None,
                                        "numero": lote.numero if lote else None,
                                        "propietario": lote.propietario if lote else None,
                                        "m2": lote.m2 if lote else None,
                                        "es_campestre": r.es_campestre,
                                        "total": 0, "detalle": {}}
            por_lote[r.lote_id]["total"] = round(por_lote[r.lote_id]["total"] + r.importe_prorrateado, 2)
            por_lote[r.lote_id]["detalle"][r.concepto] = r.importe_prorrateado
        return sorted(por_lote.values(), key=lambda x: (x["es_campestre"], x["manzana"] or 0, x["numero"] or 0))


# ─── MÓDULO 5: ESTADOS FINANCIEROS ───────────────────────────
@app.get("/api/financiero/resumen/{mes}")
def resumen_financiero(mes: str):
    with Session(engine) as s:
        lotes_vendidos = s.exec(select(Lote).where(Lote.propietario != None)).all()
        pagos_mes = s.exec(
            select(Pago).where(Pago.mes_aplicado == mes, Pago.estado == "aprobado")
        ).all()
        gastos_mes = s.exec(
            select(Gasto).where(Gasto.fecha.like(f"{mes}%"))
        ).all()
        gastos_reales = s.exec(select(GastoReal).where(GastoReal.mes == mes)).all()
        prorrateo = s.exec(select(ProrrateoPorLote).where(ProrrateoPorLote.mes == mes)).all()
        # Deuda acumulada Dicka: suma de todo el prorrateo campestre histórico
        deuda_dicka = sum(p.importe_prorrateado for p in s.exec(select(ProrrateoPorLote).where(ProrrateoPorLote.es_campestre == True)).all())
        total_ingresos = sum(p.importe for p in pagos_mes)
        total_gastos = sum(g.importe for g in gastos_mes)
        total_gastos_real = sum(g.importe for g in gastos_reales)
        total_cuotas = sum(l.cuota_cof for l in lotes_vendidos)
        total_cobrado = total_ingresos
        cartera_vencida = total_cuotas - total_cobrado
        gastos_por_concepto = {}
        for g in gastos_mes:
            gastos_por_concepto[g.tipo] = gastos_por_concepto.get(g.tipo, 0) + g.importe
        return {
            "mes": mes,
            "ingresos": {"total": total_ingresos, "cuotas_esperadas": total_cuotas,
                         "pct_cobrado": round(total_ingresos / total_cuotas * 100, 1) if total_cuotas else 0},
            "gastos": {"total": total_gastos, "total_real": total_gastos_real,
                       "por_concepto": gastos_por_concepto},
            "balance": total_ingresos - total_gastos,
            "cartera": {"total_adeudado": cartera_vencida,
                        "lotes_morosos": sum(1 for l in lotes_vendidos if not any(p.lote_id == l.id for p in pagos_mes))},
            "dicka": {"deuda_acumulada": round(deuda_dicka, 2),
                      "mes_actual": round(sum(p.importe_prorrateado for p in prorrateo if p.es_campestre), 2)},
        }


# ─── CONFIG ──────────────────────────────────────────────────
@app.get("/api/config/email")
def get_config_email(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        return {
            "smtp_from":   _get_cfg(s, "smtp_from"),
            "smtp_server": _get_cfg(s, "smtp_server", "smtp.office365.com"),
            "smtp_port":   int(_get_cfg(s, "smtp_port", "587")),
            "has_password": bool(_get_cfg(s, "smtp_password")),
        }

@app.post("/api/config/email")
def save_config_email(req: ConfigEmailReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        _set_cfg(s, "smtp_from",   req.smtp_from)
        _set_cfg(s, "smtp_server", req.smtp_server)
        _set_cfg(s, "smtp_port",   str(req.smtp_port))
        if req.smtp_password:
            _set_cfg(s, "smtp_password", req.smtp_password)
        s.commit()
    return {"ok": True}

@app.get("/api/config/plantilla")
def get_config_plantilla(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        return {
            "template": _get_cfg(s, "email_template"),
            "asunto":   _get_cfg(s, "email_asunto"),
        }

@app.post("/api/config/plantilla")
def save_config_plantilla(req: ConfigPlantillaReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        _set_cfg(s, "email_template", req.template)
        _set_cfg(s, "email_asunto",   req.asunto)
        s.commit()
    return {"ok": True}

@app.get("/api/config/plantilla-morosos")
def get_config_plantilla_morosos(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        return {
            "template": _get_cfg(s, "email_plantilla_morosos"),
            "asunto": _get_cfg(s, "email_asunto_morosos"),
        }

@app.post("/api/config/plantilla-morosos")
def save_config_plantilla_morosos(req: ConfigPlantillaReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        _set_cfg(s, "email_plantilla_morosos", req.template)
        _set_cfg(s, "email_asunto_morosos", req.asunto)
        s.commit()
    return {"ok": True}

@app.get("/api/config/auto")
def get_config_auto(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        return {
            "morosos": _get_cfg(s, "auto_morosos", "false") == "true",
            "estado":  _get_cfg(s, "auto_estado",  "false") == "true",
            "dia":     int(_get_cfg(s, "auto_dia",  "5")),
            "hora":    _get_cfg(s, "auto_hora", "09:00"),
            "activo":  _get_cfg(s, "auto_activo", "false") == "true",
        }

@app.post("/api/config/auto")
def save_config_auto(req: ConfigAutoReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        _set_cfg(s, "auto_morosos", "true" if req.morosos else "false")
        _set_cfg(s, "auto_estado",  "true" if req.estado  else "false")
        _set_cfg(s, "auto_dia",     str(req.dia))
        _set_cfg(s, "auto_hora",    req.hora)
        _set_cfg(s, "auto_activo",  "true" if req.activo  else "false")
        s.commit()
    _reload_scheduler()
    return {"ok": True}


# ─── CORREO ──────────────────────────────────────────────────
@app.post("/api/correo/enviar")
def enviar_correo(req: EnviarCorreoReq, admin: Usuario = Depends(_admin_only)):
    try:
        _send_smtp_msg(req.to, req.subject, req.body,
                       req.pdf_base64 or "", req.pdf_filename or "")
        return {"ok": True}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))

@app.post("/api/correo/prueba")
def prueba_correo(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        smtp_from = _get_cfg(s, "smtp_from")
    if not smtp_from:
        raise HTTPException(status_code=400, detail="SMTP no configurado")
    try:
        _send_smtp_msg(
            smtp_from,
            "Correo de prueba — Silvestra",
            "Este es un correo de prueba del sistema Silvestra.\n\n"
            "Si lo recibiste, la configuración SMTP es correcta.\n\n"
            "— Administración Silvestra-Canoas AC",
        )
        return {"ok": True}
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc))


# ─── CUOTAS ANUALES ──────────────────────────────────────────
def _get_cuota_para_lote(s: Session, lote: Lote, anio: int) -> Optional[float]:
    rangos = s.exec(
        select(CuotaAnual).where(CuotaAnual.anio == anio).order_by(CuotaAnual.min_m2)
    ).all()
    for r in rangos:
        tope = r.max_m2 if r.max_m2 is not None else float("inf")
        if r.min_m2 <= lote.m2 < tope:
            return r.importe
    return None

def _apply_cuotas_anuales(anio: int = None):
    if anio is None:
        anio = datetime.utcnow().year
    with Session(engine) as s:
        lotes = s.exec(select(Lote).where(Lote.activo == True)).all()
        updated = 0
        for lote in lotes:
            nueva = _get_cuota_para_lote(s, lote, anio)
            if nueva is not None and lote.cuota_cof != nueva:
                lote.cuota_cof = nueva
                s.add(lote)
                updated += 1
        s.commit()
    print(f"[Cuotas] {updated} lotes actualizados con cuotas de {anio}")
    return updated

@app.get("/api/cuotas")
def get_cuotas(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        rangos = s.exec(select(CuotaAnual).order_by(CuotaAnual.anio, CuotaAnual.min_m2)).all()
    by_year: dict = {}
    for r in rangos:
        by_year.setdefault(r.anio, []).append({
            "id": r.id, "min_m2": r.min_m2, "max_m2": r.max_m2, "importe": r.importe
        })
    return by_year

@app.post("/api/cuotas")
def crear_cuota(req: CuotaAnualReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        c = CuotaAnual(anio=req.anio, min_m2=req.min_m2, max_m2=req.max_m2, importe=req.importe)
        s.add(c); s.commit(); s.refresh(c)
        return {"id": c.id, "anio": c.anio, "min_m2": c.min_m2, "max_m2": c.max_m2, "importe": c.importe}

@app.put("/api/cuotas/{cid}")
def actualizar_cuota(cid: int, req: CuotaAnualReq, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        c = s.get(CuotaAnual, cid)
        if not c:
            raise HTTPException(404, "Rango no encontrado")
        c.anio = req.anio; c.min_m2 = req.min_m2; c.max_m2 = req.max_m2; c.importe = req.importe
        s.add(c); s.commit()
        return {"ok": True}

@app.delete("/api/cuotas/{cid}")
def eliminar_cuota(cid: int, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        c = s.get(CuotaAnual, cid)
        if not c:
            raise HTTPException(404, "Rango no encontrado")
        s.delete(c); s.commit()
        return {"ok": True}

@app.get("/api/config/deuda-campestre")
def get_deuda_campestre(admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        cfg = s.exec(select(Config).where(Config.clave == "deuda_campestre")).first()
        return {"valor": float(cfg.valor) if cfg and cfg.valor else 0.0}

@app.post("/api/config/deuda-campestre")
def set_deuda_campestre(req: dict, admin: Usuario = Depends(_admin_only)):
    with Session(engine) as s:
        cfg = s.exec(select(Config).where(Config.clave == "deuda_campestre")).first()
        if cfg:
            cfg.valor = str(req.get("valor", 0))
            s.add(cfg)
        else:
            s.add(Config(clave="deuda_campestre", valor=str(req.get("valor", 0))))
        s.commit()
    return {"ok": True}

@app.post("/api/cuotas/{anio}/aplicar")
def aplicar_cuotas(anio: int, admin: Usuario = Depends(_admin_only)):
    updated = _apply_cuotas_anuales(anio)
    return {"ok": True, "actualizados": updated}

@app.post("/api/pagos/importar")
async def importar_pagos(file: UploadFile = File(...), admin: Usuario = Depends(_admin_only)):
    import openpyxl
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Detect header row
    header_row = 0
    for i, row in enumerate(rows[:10]):
        vals = [str(v).lower() if v else "" for v in row]
        if any("manzana" in v or "mza" in v for v in vals) and any("lote" in v for v in vals):
            header_row = i
            break
    headers = [str(v).lower().strip() if v else "" for v in rows[header_row]]

    col = {}
    for i, h in enumerate(headers):
        if "fecha" in h:                        col["fecha"] = i
        elif "manzana" in h or h == "mza":      col["mza"] = i
        elif h == "lote" or h == "lote_num":    col["lote"] = i
        elif "concepto" in h or "tipo" in h:    col["concepto"] = i
        elif "importe" in h or "monto" in h or "cuota" in h: col["importe"] = i
        elif "referen" in h or "folio" in h or h == "ref":   col["ref"] = i

    importados, errores = 0, []
    with Session(engine) as s:
        for idx, row in enumerate(rows[header_row + 1:], start=2):
            if not row or not any(row): continue
            try:
                raw_fecha = row[col["fecha"]] if "fecha" in col else None
                if not raw_fecha: continue
                if hasattr(raw_fecha, "date"):
                    fecha = raw_fecha.date()
                else:
                    fecha = date.fromisoformat(str(raw_fecha)[:10])

                mza = int(row[col["mza"]]) if "mza" in col else None
                lote_num = int(row[col["lote"]]) if "lote" in col else None
                if not mza or not lote_num: continue

                importe = float(row[col["importe"]]) if "importe" in col else 0
                if importe <= 0: continue

                concepto = str(row[col["concepto"]] or "COF").strip() if "concepto" in col else "COF"
                ref = str(row[col["ref"]] or "").strip() if "ref" in col else None

                lote = s.exec(select(Lote).where(Lote.manzana == mza, Lote.numero == lote_num)).first()
                if not lote:
                    errores.append(f"Fila {idx}: Mza {mza} L{lote_num} no encontrado")
                    continue

                s.add(Pago(
                    lote_id=lote.id,
                    fecha_pago=fecha,
                    importe=importe,
                    mes_aplicado=str(fecha)[:7],
                    concepto=concepto,
                    referencia=ref or None,
                    estado="por_aprobar",
                ))
                importados += 1
            except Exception as e:
                errores.append(f"Fila {idx}: {e}")
        s.commit()

    return {"importados": importados, "errores": errores}


@app.post("/api/historial/importar")
async def importar_historial(file: UploadFile = File(...), admin: Usuario = Depends(_admin_only)):
    """Lee todas las hojas Edo.Cta.M* del Excel y crea Pago + Descuento históricos."""
    import openpyxl, re
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    hoy = date.today()
    pagos_creados = descuentos_creados = lotes_actualizados = 0
    errores = []

    patron = re.compile(r'Edo\.Cta\.M(\d+)\s+L(\d+)', re.IGNORECASE)

    with Session(engine) as s:
        for nombre_hoja in wb.sheetnames:
            m = patron.match(nombre_hoja)
            if not m:
                continue
            mza, lote_num = int(m.group(1)), int(m.group(2))
            lote = s.exec(select(Lote).where(Lote.manzana == mza, Lote.numero == lote_num)).first()
            if not lote:
                continue

            ws = wb[nombre_hoja]
            rows = list(ws.iter_rows(values_only=True))

            # Actualizar fecha_escrituracion desde Row 6 (idx 5), Col 2
            try:
                fecha_esc_raw = rows[5][2] if len(rows) > 5 else None
                if fecha_esc_raw and hasattr(fecha_esc_raw, 'date'):
                    lote.fecha_escrituracion = fecha_esc_raw.date()
                    s.add(lote)
                    lotes_actualizados += 1
            except Exception:
                pass

            # Datos mensuales desde fila 19 (idx 18)
            for row in rows[18:]:
                if not row or len(row) < 7:
                    continue
                mes_val = row[0]
                if not hasattr(mes_val, 'year'):
                    continue
                mes_date = date(mes_val.year, mes_val.month, 1)
                if mes_date > hoy:
                    continue
                mes_str = mes_date.strftime('%Y-%m')

                # COF payment
                pago_cof = row[3] if len(row) > 3 else None
                desc_cof  = row[4] if len(row) > 4 else None
                fecha_cof  = row[5] if len(row) > 5 else None

                if pago_cof and float(pago_cof) > 0:
                    existe = s.exec(select(Pago).where(
                        Pago.lote_id == lote.id, Pago.mes_aplicado == mes_str,
                        Pago.concepto == "COF", Pago.estado == "aprobado",
                    )).first()
                    if not existe:
                        fp = fecha_cof.date() if fecha_cof and hasattr(fecha_cof, 'date') else mes_date
                        s.add(Pago(lote_id=lote.id, fecha_pago=fp, importe=float(pago_cof),
                                   mes_aplicado=mes_str, concepto="COF", estado="aprobado",
                                   notas="Histórico importado"))
                        pagos_creados += 1

                if desc_cof and float(desc_cof) > 0:
                    existe_desc = s.exec(select(Descuento).where(
                        Descuento.lote_id == lote.id, Descuento.concepto == "COF",
                        Descuento.anio == mes_date.year,
                    )).first()
                    if not existe_desc:
                        s.add(Descuento(lote_id=lote.id, tipo="descuento", concepto="COF",
                                        importe=float(desc_cof), anio=mes_date.year,
                                        fecha=mes_date, notas="Histórico importado"))
                        descuentos_creados += 1

                # COV payment
                pago_cov  = row[12] if len(row) > 12 else None
                fecha_cov  = row[13] if len(row) > 13 else None

                if pago_cov and float(pago_cov) > 0:
                    existe_cov = s.exec(select(Pago).where(
                        Pago.lote_id == lote.id, Pago.mes_aplicado == mes_str,
                        Pago.concepto == "COV", Pago.estado == "aprobado",
                    )).first()
                    if not existe_cov:
                        fp = fecha_cov.date() if fecha_cov and hasattr(fecha_cov, 'date') else mes_date
                        s.add(Pago(lote_id=lote.id, fecha_pago=fp, importe=float(pago_cov),
                                   mes_aplicado=mes_str, concepto="COV", estado="aprobado",
                                   notas="Histórico importado"))
                        pagos_creados += 1

        s.commit()

    return {"ok": True, "pagos_creados": pagos_creados,
            "descuentos_creados": descuentos_creados,
            "lotes_actualizados": lotes_actualizados,
            "errores": errores[:20]}

@app.post("/api/admin/importar_silvestra_json")
async def importar_silvestra_json(admin: Usuario = Depends(_admin_only)):
    """Migración: importa silvestra_data.json completo a tablas Pago, Descuento y LecturaAgua."""
    import json as _json
    json_path = os.path.join(os.path.dirname(__file__), "static", "silvestra_data.json")
    if not os.path.exists(json_path):
        raise HTTPException(404, "silvestra_data.json no encontrado")
    with open(json_path) as f:
        sd = _json.load(f)

    pagos_creados = descuentos_creados = lotes_actualizados = lecturas_creadas = 0

    with Session(engine) as s:
        lotes_db = {(l.manzana, l.numero): l for l in s.exec(select(Lote)).all()}

        for ld in sd:
            lote = lotes_db.get((ld.get("mza"), ld.get("lote")))
            if not lote:
                continue

            # Actualizar fecha de escrituración
            if ld.get("escrituracion"):
                try:
                    lote.fecha_escrituracion = date.fromisoformat(ld["escrituracion"])
                    s.add(lote); lotes_actualizados += 1
                except: pass

            for m in ld.get("meses", []):
                mes_str = (m.get("mes") or "")[:7]   # "2020-10-01" → "2020-10"
                if not mes_str or len(mes_str) < 7:
                    continue
                try:
                    mes_date = date.fromisoformat(mes_str + "-01")
                except:
                    continue

                def _fp(m_dict):
                    raw = m_dict.get("fpago")
                    if raw:
                        try: return date.fromisoformat(str(raw)[:10])
                        except: pass
                    return mes_date

                pago_cof = float(m.get("pagoCof") or 0)
                if pago_cof > 0:
                    if not s.exec(select(Pago).where(Pago.lote_id==lote.id,Pago.mes_aplicado==mes_str,Pago.concepto=="COF",Pago.estado=="aprobado")).first():
                        s.add(Pago(lote_id=lote.id,fecha_pago=_fp(m),importe=pago_cof,mes_aplicado=mes_str,concepto="COF",estado="aprobado",notas="silvestra_json"))
                        pagos_creados += 1

                desc_cof = float(m.get("desc") or 0)
                if desc_cof > 0:
                    if not s.exec(select(Descuento).where(Descuento.lote_id==lote.id,Descuento.concepto=="COF",Descuento.anio==mes_date.year)).first():
                        s.add(Descuento(lote_id=lote.id,tipo="descuento",concepto="COF",importe=desc_cof,anio=mes_date.year,fecha=mes_date,notas="silvestra_json"))
                        descuentos_creados += 1

                pago_cov = float(m.get("pagoCov") or 0)
                if pago_cov > 0:
                    if not s.exec(select(Pago).where(Pago.lote_id==lote.id,Pago.mes_aplicado==mes_str,Pago.concepto=="COV",Pago.estado=="aprobado")).first():
                        s.add(Pago(lote_id=lote.id,fecha_pago=_fp(m),importe=pago_cov,mes_aplicado=mes_str,concepto="COV",estado="aprobado",notas="silvestra_json"))
                        pagos_creados += 1

                if float(m.get("consumo") or 0) > 0 or float(m.get("cov") or 0) > 0:
                    if not s.exec(select(LecturaAgua).where(LecturaAgua.lote_id==lote.id,LecturaAgua.mes==mes_str)).first():
                        s.add(LecturaAgua(lote_id=lote.id,mes=mes_str,
                            lectura_anterior=float(m.get("vIni") or 0),lectura_actual=float(m.get("vFin") or 0),
                            consumo_m3=float(m.get("consumo") or 0),tarifa_por_m3=float(ld.get("tarifaCOV") or 15.0),
                            importe=float(m.get("cov") or 0)))
                        lecturas_creadas += 1

        s.commit()

    return {"ok": True, "pagos_creados": pagos_creados, "descuentos_creados": descuentos_creados,
            "lotes_actualizados": lotes_actualizados, "lecturas_creadas": lecturas_creadas}


@app.post("/api/admin/regenerar_desde_excel")
async def regenerar_desde_excel(file: UploadFile = File(...), admin: Usuario = Depends(_admin_only)):
    """Re-genera silvestra_data.json desde las hojas Edo.Cta.* y reimporta Chequera (limpia primero)."""
    import openpyxl, json as _json
    from sqlalchemy import text as _text2

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    lotes_data = []

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith('Edo.Cta.'):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 14:
            continue

        nombre = str(rows[5][2] or '').strip() if len(rows[5]) > 2 and rows[5][2] else ''
        if not nombre:
            continue

        def _fv(r, col, default=0.0):
            v = r[col] if len(r) > col else None
            if v is None: return default
            try: return float(v)
            except: return default

        mza = int(_fv(rows[6], 2, 0))
        lote_num = int(_fv(rows[7], 2, 0))
        m2 = _fv(rows[6], 9, 0.0)
        cuota_fija = _fv(rows[5], 13, 0.0)
        tarifa_cov = _fv(rows[6], 13, 5.0)
        saldo_cof = _fv(rows[13], 6, 0.0)

        escrit_raw = rows[5][9] if len(rows[5]) > 9 else None
        escrituracion = escrit_raw.date().isoformat() if hasattr(escrit_raw, 'date') else ''

        meses = []
        for row in rows[18:]:
            if not row or len(row) < 1 or not row[0] or not hasattr(row[0], 'year'):
                continue
            meses.append({
                'mes': row[0].date().isoformat(),
                'extra': _fv(row, 1),
                'cof': _fv(row, 2),
                'pagoCof': _fv(row, 3),
                'desc': _fv(row, 4),
                'fpago': row[5].date().isoformat() if len(row) > 5 and hasattr(row[5], 'date') else '',
                'saldoMes': _fv(row, 6),
                'vIni': _fv(row, 7),
                'vFin': _fv(row, 8),
                'consumo': _fv(row, 9),
                'cov': _fv(row, 10),
                'instal': _fv(row, 11),
                'pagoCov': _fv(row, 12),
            })

        lotes_data.append({
            'sheet': sheet_name,
            'mza': mza, 'lote': lote_num, 'nombre': nombre,
            'm2': m2, 'escrituracion': escrituracion,
            'cuotaFija': cuota_fija, 'tarifaCOV': tarifa_cov,
            'totales': {
                'extra': round(sum(m['extra'] for m in meses), 2),
                'cof': round(sum(m['cof'] for m in meses), 2),
                'pagoCof': round(sum(m['pagoCof'] for m in meses), 2),
                'desc': round(sum(m['desc'] for m in meses), 2),
                'saldoCof': round(saldo_cof, 2),
                'cov': round(sum(m['cov'] for m in meses), 2),
                'instal': round(sum(m['instal'] for m in meses), 2),
                'pagoCov': round(sum(m['pagoCov'] for m in meses), 2),
            },
            'meses': meses
        })

    lotes_data.sort(key=lambda x: (x['mza'], x['lote']))
    json_path = os.path.join(os.path.dirname(__file__), 'static', 'silvestra_data.json')
    with open(json_path, 'w', encoding='utf-8') as f:
        _json.dump(lotes_data, f, ensure_ascii=False, separators=(',', ':'))

    # Upsert lotes into DB (manzana, lote, propietario, m2, cuota, escrituracion)
    with Session(engine) as s:
        for ld in lotes_data:
            lote_row = s.exec(select(Lote).where(
                Lote.manzana == ld['mza'], Lote.numero == ld['lote']
            )).first()
            escrit = None
            if ld['escrituracion']:
                try: escrit = date.fromisoformat(ld['escrituracion'])
                except: pass
            if lote_row:
                lote_row.propietario = ld['nombre']
                lote_row.m2 = ld['m2']
                lote_row.cuota_cof = ld['cuotaFija']
                lote_row.fecha_escrituracion = escrit
                lote_row.activo = True
                s.add(lote_row)
            else:
                s.add(Lote(
                    manzana=ld['mza'], numero=ld['lote'],
                    propietario=ld['nombre'], m2=ld['m2'],
                    cuota_cof=ld['cuotaFija'], paseo='',
                    fecha_escrituracion=escrit, activo=True
                ))
        s.commit()

    # Limpiar y reimportar Chequera
    with engine.connect() as conn:
        conn.execute(_text2("DELETE FROM movimientobancario"))
        conn.commit()

    ws_cheq = wb['Chequera']
    importados = 0
    with Session(engine) as s:
        for row in list(ws_cheq.iter_rows(values_only=True))[5:]:
            if not row or len(row) < 3: continue
            fecha_raw = row[2]
            if not fecha_raw: continue
            if hasattr(fecha_raw, 'date'):
                fecha = fecha_raw.date()
            else:
                try: fecha = date.fromisoformat(str(fecha_raw)[:10])
                except:
                    try: fecha = date(1899, 12, 30) + timedelta(days=int(float(str(fecha_raw))))
                    except: continue

            def _sv(r, c): return str(r[c] or '').strip() if len(r) > c and r[c] else ''
            def _nv(r, c): return float(r[c] or 0) if len(r) > c and r[c] else 0.0

            nombre_v = _sv(row, 3); concepto_v = _sv(row, 4)
            s.add(MovimientoBancario(
                fecha=fecha,
                descripcion=nombre_v or concepto_v or 'Sin descripción',
                nombre=nombre_v, concepto=concepto_v,
                importe=float(row[5]) if len(row) > 5 and row[5] else None,
                cargo=_nv(row, 10), abono=_nv(row, 11),
                referencia=_sv(row, 1) or None,
                tipo='sin_clasificar',
                iva_ret=_nv(row, 6) or _nv(row, 8),
                isr_ret=_nv(row, 7),
                tiene_iva_ret=bool(_nv(row, 6) or _nv(row, 8)),
                tiene_isr_ret=bool(_nv(row, 7)),
            ))
            importados += 1
        s.commit()

    return {
        "ok": True,
        "lotes_actualizados": len(lotes_data),
        "meses_total": sum(len(l['meses']) for l in lotes_data),
        "chequera_importados": importados,
        "lotes_en_bd": len(lotes_data)
    }
