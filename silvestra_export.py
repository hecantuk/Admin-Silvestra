"""
Exportación del libro Excel "estado de cuenta Silvestra" desde la base de datos,
replicando el formato y las FÓRMULAS del archivo original
(Estados_de_cuenta_colonos_al_31_de_mayo_2026.xlsm).

Genera una hoja `Edo.Cta.M## L##` por lote (con fórmulas vivas de saldo, consumo,
COV y totales), más `Chequera`, `Concentrado COF` y `Concentrado COV`.

La cuota mensual (columna C) y la cuota fija del encabezado (N6) se escriben como
VALOR real de la BD (no como LOOKUP), para no depender de la pestaña Cuotas; todas
las celdas de cálculo (saldos, consumo, COV, totales) sí son fórmulas, de modo que
si el usuario edita un pago o una lectura en el Excel, los saldos se recalculan.
"""
import io
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ── Formatos de número (idénticos al original) ──────────────────────────
FMT_COF   = r'_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'          # COF: contable, 0 dec
FMT_COV   = r'_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'   # COV: contable, 2 dec
FMT_CURR0 = r'"$"#,##0'
FMT_CURR2 = r'"$"#,##0.00'
FMT_MES   = 'mmm-yy'
FMT_FECHA = 'd-mmm-yy'

# Estilos
F_TITLE = Font(name='Calibri', size=12, bold=True)
F_SUB   = Font(name='Calibri', size=10, bold=True, color='666666')
F_BOLD  = Font(name='Calibri', size=10, bold=True)
F_HEAD  = Font(name='Calibri', size=9,  bold=True, color='FFFFFF')
F_NORM  = Font(name='Calibri', size=10)
FILL_HEAD = PatternFill('solid', fgColor='2F5233')   # verde Silvestra
FILL_SEC  = PatternFill('solid', fgColor='E8EEE9')
FILL_TOT  = PatternFill('solid', fgColor='F2F2F2')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT  = Alignment(horizontal='right')
_thin = Side(style='thin', color='D0D0D0')
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

COL_W = {'A':11,'B':14,'C':10,'D':11,'E':11,'F':13,'G':12,'H':10,'I':10,
         'J':11,'K':12,'L':16,'M':11,'N':13,'O':12}

CONTACTO = ('Para cualquier duda y/o aclaración al respecto, puede comunicarse al '
            'teléfono: (81) 8865-7770 ext. 7741 con Ana Sena, correo electrónico: '
            'sena.ana@dicka.com.mx')


def _mes_date(mes_key: str) -> date:
    y, m = mes_key.split('-')[:2]
    return date(int(y), int(m), 1)


def _set(ws, coord, value, font=None, fmt=None, align=None, fill=None):
    c = ws[coord]
    c.value = value
    if font:  c.font = font
    if fmt:   c.number_format = fmt
    if align: c.alignment = align
    if fill:  c.fill = fill
    return c


def build_edocta_ws(wb, lote, meses, tarifa):
    """Crea una hoja Edo.Cta.M## L## con el formato y fórmulas del original.
    `lote`: dict con mza, lote, nombre, m2, cuotaFija, escrituracion.
    `meses`: lista de dicts (mes, extra, cof, pagoCof, desc, fpago, vIni, vFin,
             consumo, cov, instal, pagoCov)."""
    name = f"Edo.Cta.M{lote['mza']} L{int(lote['lote']):02d}"
    ws = wb.create_sheet(title=name[:31])

    for col, w in COL_W.items():
        ws.column_dimensions[col].width = w

    # ── Encabezado ──
    _set(ws, 'A1', 'Asociación de Colonos Silvestra - Canoas, A.C.', F_TITLE, align=CENTER)
    _set(ws, 'A2', 'Dicka Campestre, S.A. de C.V.', F_SUB, align=CENTER)
    _set(ws, 'A4', 'Estado de cuenta', F_BOLD)
    _set(ws, 'J5', 'Información generada a la fecha de hoy:', F_BOLD)
    _set(ws, 'N5', datetime.now(), F_BOLD, fmt='d-mmm-yy')

    _set(ws, 'A6', 'Asociado:', F_BOLD)
    _set(ws, 'C6', lote['nombre'], F_BOLD)
    _set(ws, 'G6', 'Fecha de escrituración:', F_BOLD)
    esc = lote.get('escrituracion')
    if esc:
        try: esc = date.fromisoformat(str(esc)[:10])
        except Exception: pass
    _set(ws, 'J6', esc or '—', F_BOLD, fmt=FMT_FECHA)
    _set(ws, 'L6', 'Cuota Fija:', F_BOLD)
    _set(ws, 'N6', lote.get('cuotaFija') or 0, F_BOLD, fmt=FMT_CURR0)

    _set(ws, 'A7', 'Manzana:', F_BOLD)
    _set(ws, 'C7', int(lote['mza']), F_BOLD)
    _set(ws, 'G7', 'M2:', F_BOLD)
    _set(ws, 'J7', float(lote.get('m2') or 0), F_BOLD, fmt='#,##0.00')
    _set(ws, 'L7', 'Valor cuota ordinaria variable:', F_BOLD)
    _set(ws, 'N7', float(tarifa or 5), F_BOLD, fmt=FMT_CURR2)

    _set(ws, 'A8', 'Lote:', F_BOLD)
    _set(ws, 'C8', int(lote['lote']), F_BOLD)

    # ── Resumen (filas 10-14) ──
    _set(ws, 'A10', 'Conceptos', F_HEAD, align=CENTER, fill=FILL_HEAD)
    _set(ws, 'F10', 'Cuota Ordinaria Fija de mantenimiento (COF)', F_HEAD, align=CENTER, fill=FILL_HEAD)
    _set(ws, 'J10', 'Administración y mantenimiento de redes de agua (COV)', F_HEAD, align=CENTER, fill=FILL_HEAD)
    _set(ws, 'N10', 'Totales', F_HEAD, align=CENTER, fill=FILL_HEAD)

    n = len(meses)
    first, last = 19, 18 + n          # filas de datos
    tot_row = last + 1                # fila de totales
    has = n > 0

    def rng(col): return f"{col}{first}:{col}{last}"

    _set(ws, 'A11', 'Monto total de cuotas', F_BOLD)
    _set(ws, 'G11', (f"=SUM({rng('B')})+SUM({rng('C')})" if has else 0), F_BOLD, fmt=FMT_CURR0)
    _set(ws, 'L11', (f"=SUM({rng('K')})+SUM({rng('L')})" if has else 0), F_BOLD, fmt=FMT_CURR2)
    _set(ws, 'N11', ("=G11+L11" if has else 0), F_BOLD, fmt=FMT_CURR2)

    _set(ws, 'A12', 'Monto total de pagos', F_BOLD)
    _set(ws, 'G12', (f"=SUM({rng('D')})" if has else 0), F_BOLD, fmt=FMT_CURR0)
    _set(ws, 'L12', (f"=SUM({rng('M')})" if has else 0), F_BOLD, fmt=FMT_CURR2)
    _set(ws, 'N12', ("=G12+L12" if has else 0), F_BOLD, fmt=FMT_CURR2)

    _set(ws, 'A13', 'Descuentos', F_BOLD)
    _set(ws, 'G13', (f"=SUM({rng('E')})" if has else 0), F_BOLD, fmt=FMT_CURR0)

    _set(ws, 'A14', 'Saldo pendiente de liquidar', F_BOLD)
    _set(ws, 'G14', (f"=SUM({rng('G')})" if has else 0), F_BOLD, fmt=FMT_CURR0)
    _set(ws, 'L14', ("=L11-L12" if has else 0), F_BOLD, fmt=FMT_CURR2)
    _set(ws, 'N14', ("=G14+L14" if has else 0), F_BOLD, fmt=FMT_CURR2)

    # ── Detalle (encabezados fila 18) ──
    _set(ws, 'A16', 'Detalle', F_BOLD)
    _set(ws, 'A17', 'Cuota Ordinaria Fija de mantenimiento (COF)', F_BOLD, fill=FILL_SEC)
    _set(ws, 'H17', 'Administración y mantenimiento de redes de agua (COV)', F_BOLD, fill=FILL_SEC)
    headers = ['Mes', 'Cuota extraordinaria única', 'COF', 'Pagos', 'Descuento',
               'Fecha de pago', 'Saldo', 'Vol. Inicial', 'Vol. Final',
               'Consumo (m3)', 'COV', 'Instalación de medidor volumétrico.',
               'Pagos', 'Fecha de pago', 'Saldo']
    for i, h in enumerate(headers):
        _set(ws, f"{get_column_letter(i+1)}18", h, F_HEAD, align=CENTER, fill=FILL_HEAD)

    # ── Filas de detalle (valores en columnas manuales, fórmulas en cálculo) ──
    for idx, m in enumerate(meses):
        r = first + idx
        has_lect = (m.get('vFin') is not None) and (m.get('vIni') is not None)
        _set(ws, f'A{r}', _mes_date(m['mes']), fmt=FMT_MES)
        _set(ws, f'B{r}', m.get('extra') or 0, fmt=FMT_COF)
        _set(ws, f'C{r}', m.get('cof') or 0, fmt=FMT_COF)
        _set(ws, f'D{r}', m.get('pagoCof') or 0, fmt=FMT_COF)
        _set(ws, f'E{r}', m.get('desc') or 0, fmt=FMT_COF)
        fp = m.get('fpago')
        if fp:
            try: fp = date.fromisoformat(str(fp)[:10])
            except Exception: pass
        _set(ws, f'F{r}', fp or None, fmt=FMT_FECHA)
        _set(ws, f'G{r}', f"=B{r}+C{r}-D{r}-E{r}", F_BOLD, fmt=FMT_COF)   # Saldo COF mes
        _set(ws, f'H{r}', m.get('vIni') if has_lect else None, fmt='#,##0')
        _set(ws, f'I{r}', m.get('vFin') if has_lect else None, fmt='#,##0')
        if has_lect:
            _set(ws, f'J{r}', f"=I{r}-H{r}", fmt='#,##0')                 # Consumo
            _set(ws, f'K{r}', f"=J{r}*$N$7", fmt=FMT_COV)                 # COV = consumo*tarifa
        else:
            _set(ws, f'J{r}', 0, fmt='#,##0')
            _set(ws, f'K{r}', m.get('cov') or 0, fmt=FMT_COV)
        _set(ws, f'L{r}', m.get('instal') or 0, fmt=FMT_COV)
        _set(ws, f'M{r}', m.get('pagoCov') or 0, fmt=FMT_COV)
        _set(ws, f'N{r}', None, fmt=FMT_FECHA)                           # fecha pago COV (no en BD)
        _set(ws, f'O{r}', f"=K{r}+L{r}-M{r}", fmt=FMT_COV)               # Saldo COV mes

    # ── Fila de totales ──
    if has:
        _set(ws, f'A{tot_row}', 'Totales', F_BOLD, fill=FILL_TOT)
        for col, fmt in (('B', FMT_COF), ('C', FMT_COF), ('D', FMT_COF), ('E', FMT_COF),
                         ('G', FMT_COF), ('K', FMT_COV), ('L', FMT_COV), ('M', FMT_COV),
                         ('O', FMT_COV)):
            _set(ws, f'{col}{tot_row}', f"=SUM({col}{first}:{col}{last})", F_BOLD, fmt=fmt, fill=FILL_TOT)

    # ── Pie ──
    f = tot_row + 2
    _set(ws, f'A{f}', CONTACTO, F_NORM)
    _set(ws, f'A{f+1}', 'Los pagos se realizan por transferencia o depósito a la cuenta:', F_NORM)
    _set(ws, f'A{f+2}', 'Banco:', F_BOLD); _set(ws, f'E{f+2}', 'Banregio')
    _set(ws, f'A{f+3}', 'Beneficiario:', F_BOLD); _set(ws, f'E{f+3}', 'Asociación de Colonos Silvestra-Canoas AC')
    _set(ws, f'A{f+4}', 'Cuenta:', F_BOLD); _set(ws, f'E{f+4}', '001166030010')

    # ── Notas (anotaciones libres: tarjetas de acceso, condonaciones, etc.) ──
    notas = (lote.get('notas') or '').strip()
    if notas:
        nr = f + 6
        _set(ws, f'A{nr}', 'Notas:', F_BOLD, fill=FILL_SEC)
        for i, linea in enumerate(notas.split('\n')):
            _set(ws, f'A{nr+1+i}', linea, F_NORM)

    # Merges del encabezado
    for mr in ('A1:O1', 'A2:O2', 'A4:O4', 'A16:O16', 'A17:G17', 'H17:O17',
               'F10:I10', 'J10:M10', 'N10:O10', 'G11:H11', 'N11:O11',
               'G12:H12', 'N12:O12', 'G14:H14', 'N14:O14'):
        try: ws.merge_cells(mr)
        except Exception: pass
    return ws


def build_chequera_ws(wb, movimientos):
    ws = wb.create_sheet('Chequera')
    for col, w in {'A':10,'B':16,'C':12,'D':26,'E':30,'F':14,'G':12,'H':12,'I':12,'J':12,'K':14,'L':14}.items():
        ws.column_dimensions[col].width = w
    _set(ws, 'A1', 'Chequera — Asociación de Colonos Silvestra', F_TITLE)
    headers = ['#', 'Referencia', 'Fecha', 'Nombre', 'Concepto', 'Importe',
               'IVA ret.', 'ISR ret.', 'IVA ret.', 'Cargo', 'Abono', 'Saldo']
    for i, h in enumerate(headers):
        _set(ws, f"{get_column_letter(i+1)}5", h, F_HEAD, align=CENTER, fill=FILL_HEAD)
    r = 6
    for i, m in enumerate(movimientos):
        _set(ws, f'A{r}', i + 1)
        _set(ws, f'B{r}', m.get('referencia') or '')
        fch = m.get('fecha')
        if fch:
            try: fch = date.fromisoformat(str(fch)[:10])
            except Exception: pass
        _set(ws, f'C{r}', fch, fmt=FMT_FECHA)
        _set(ws, f'D{r}', m.get('nombre') or '')
        _set(ws, f'E{r}', m.get('concepto') or m.get('descripcion') or '')
        _set(ws, f'F{r}', m.get('importe'), fmt=FMT_CURR2)
        _set(ws, f'G{r}', m.get('iva_ret') or 0, fmt=FMT_CURR2)
        _set(ws, f'H{r}', m.get('isr_ret') or 0, fmt=FMT_CURR2)
        _set(ws, f'J{r}', m.get('cargo') or 0, fmt=FMT_CURR2)
        _set(ws, f'K{r}', m.get('abono') or 0, fmt=FMT_CURR2)
        # Saldo corrido (fórmula)
        if r == 6:
            _set(ws, f'L{r}', f"=K{r}-J{r}", fmt=FMT_CURR2)
        else:
            _set(ws, f'L{r}', f"=L{r-1}+K{r}-J{r}", fmt=FMT_CURR2)
        r += 1
    return ws


def build_concentrado_ws(wb, titulo, lotes, key):
    """Concentrado COF o COV: una fila por lote, una columna por año + total.
    `key`='cof' usa saldoCof y los meses cof/extra/pagoCof; 'cov' usa saldoCov."""
    ws = wb.create_sheet(titulo)
    anios = set()
    for l in lotes:
        for m in l['meses']:
            anios.add(m['mes'][:4])
    anios = sorted(anios)
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 6
    ws.column_dimensions['C'].width = 30
    for i in range(len(anios) + 1):
        ws.column_dimensions[get_column_letter(4 + i)].width = 12
    _set(ws, 'A1', titulo, F_TITLE)
    head = ['Mza', 'Lote', 'Asociado'] + anios + ['Saldo total']
    for i, h in enumerate(head):
        _set(ws, f"{get_column_letter(i+1)}3", h, F_HEAD, align=CENTER, fill=FILL_HEAD)
    r = 4
    for l in lotes:
        _set(ws, f'A{r}', l['mza']); _set(ws, f'B{r}', l['lote'])
        _set(ws, f'C{r}', l['nombre'])
        por_anio = {a: 0.0 for a in anios}
        for m in l['meses']:
            a = m['mes'][:4]
            if key == 'cof':
                por_anio[a] += (m.get('cof') or 0) + (m.get('extra') or 0) - (m.get('pagoCof') or 0) - (m.get('desc') or 0)
            else:
                por_anio[a] += (m.get('cov') or 0) + (m.get('instal') or 0) - (m.get('pagoCov') or 0)
        for i, a in enumerate(anios):
            _set(ws, f"{get_column_letter(4+i)}{r}", round(por_anio[a], 2),
                 fmt=(FMT_CURR0 if key == 'cof' else FMT_CURR2))
        saldo = l['saldoCof'] if key == 'cof' else l['saldoCov']
        _set(ws, f"{get_column_letter(4+len(anios))}{r}", round(saldo or 0, 2),
             F_BOLD, fmt=(FMT_CURR0 if key == 'cof' else FMT_CURR2))
        r += 1
    return ws


def construir_libro(lotes_full, movimientos):
    """lotes_full: lista de dicts del concentrado (mza, lote, nombre, m2, cuotaFija,
    escrituracion, saldoCof, saldoCov, meses, tarifaCOV). Devuelve bytes .xlsx."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # quitar hoja por defecto

    # Orden como el original: Concentrados primero, luego lotes, luego Chequera
    build_concentrado_ws(wb, 'Concentrado COF', lotes_full, 'cof')
    build_concentrado_ws(wb, 'Concentrado COV', lotes_full, 'cov')

    lotes_ord = sorted(lotes_full, key=lambda x: (x['mza'], x['lote']))
    for l in lotes_ord:
        meses = [m for m in l['meses']
                 if (m.get('cof') or m.get('cov') or m.get('pagoCof') or m.get('extra')
                     or m.get('pagoCov') or m.get('instal'))]
        build_edocta_ws(wb, l, meses, l.get('tarifaCOV', 5))

    build_chequera_ws(wb, movimientos)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
